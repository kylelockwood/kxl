#! python3
"""
Return excel data based on criteria in a variety of formats
"""
import openpyxl
import datetime as dt

class data:
    def __init__(self, 
                filePath: str, 
                sheet='Sheet1', 
                row_range=[1],
                col_range=[1],
                delimiter=' ',
                skip_none=True,
                alerts=True):
        self.alerts = alerts
        self.wb = self.__load_workbook__(filePath)
        self.sheet = self.wb[sheet]
        if len(row_range) == 1:
            self.row_range = row_range
        else:
            self.row_range = row_range[0], row_range[1] + 1
        if len(col_range) == 1:
            self.col_range = col_range
        else:
            self.col_range = col_range[0], col_range[1] + 1 
        self.delimiter = delimiter
        self.value = self.get_value(self.row_range[0], self.col_range[0])
        self.skip_none = skip_none

    def __load_workbook__(self, path):
        """ Loads the excel workbook for reading """
        if self.alerts is True:
            print(f'Reading \'{path}\'... ', flush=True, end='')
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if self.alerts is True:
            print('Done')
        return workbook

    def dict_list(self, keys='columns', key_names=[], key_index=1):
        """ 
        Returns a dict of keys with list of string data from specified ranges 
        If keys are not provided, you may pass the location of those keys in the sheet
        Location of keys is chosen by 'columns or 'rows' and the key_index of that choice
        i.e. keys='rows', key_index=1 means the key names are in column 1 on each row -
        the cooresponding data will be returned as a list from the remaining data in col_range 
        """

        # Create keys if they aren't provided
        if not key_names:
            if keys is 'columns':
                key_range = self.col_range
            else:
                key_range = self.row_range
            for f in range(key_range[0], key_range[1]):
                if keys is 'columns':
                    r = key_index
                    c = f
                else:
                    r = f
                    c = key_index
                keyname = self.sheet.cell(row=r, column=c).value
                key_names.append(keyname)
        
        # Create dict of lists
        if keys is 'columns':
            first_range = self.row_range
            second_range = self.col_range
        else:
            first_range = self.col_range
            second_range = self.row_range
        dataDict = {}
        for i in range(len(key_names)):
            val = ""
            valuelist = []
            for f in range(first_range[0], first_range[1]):
                if keys is 'columns':
                    r = f
                    c = second_range[0] + i
                else:
                    r = second_range[0] + i
                    c = f
                
                val = self.sheet.cell(row=r, column=c).value
                if self.skip_none is True and val is None:
                    continue
                valuelist.append(val)
            dataDict[key_names[i]] = valuelist
        return dataDict

    def list_of(self, list_type='string',
                    row_range=None, 
                    col_range=None, 
                    keys=None,
                    key_names=[],
                    key_index=1, 
                    date_format='%m/%d/%Y %H:%M:%S',
                    delimiter=None,
                    skip_none=None,
                    alerts=None):

        """
        Returns a list of list_type data from passed sheet.  
        If skip_none is True (default), blank lines are skipped
        list_type 'string':
            If the final list contains only one item, that item will be returned as a string
            Keys determines the start of the list item.  i.e. if keys are 'column', the list
            will be populated with concatenated strings from each column
        list_type 'list:
            If the final list contains lists with single items, a single list will be created and sent instead
        list_type 'dict':
            See def dict_list
        """
        
        # Set default values
        if alerts is None:
            alerts = self.alerts
        if alerts is True:
            print(f'Collecting data from worksheet \'{self.sheet.title}\'... ', end='', flush=True)
        if skip_none is None:
            skip_none = self.skip_none
        if row_range is None:
            row_range = self.row_range
        if col_range is None:
            col_range = self.col_range
        if len(row_range) == 1:
            row_range = row_range[0], row_range[0] + 1
        if len(col_range) == 1:
            col_range = col_range[0], col_range[0] + 1
            # No delimiter needed if single column
            if delimiter is None:
                delimiter = ''
        if list_type is 'dict' and keys is None:
            keys = 'columns'
        if keys is None:
            keys = 'rows'
        if list_type is 'dict':
            return self.dict_list(keys, key_names, key_index)
        if delimiter is None:
            delimiter = self.delimiter
        
        # Read data and create list
        dataList = []
        if keys is 'rows':
            first_range = row_range
            second_range = col_range
        else:
            first_range = col_range
            second_range = row_range
        for f in range(first_range[0], first_range[1]):
            if list_type is 'string':
                dataCol = ''
            elif list_type is 'list':
                dataCol = []
            for s in range(second_range[0], second_range[1]):
                cellData = self.sheet.cell(row=f, column=s).value
                if skip_none is True and cellData is None:
                    continue
                else:
                    if list_type is 'string':
                        if isinstance(cellData, dt.datetime):
                            cellData = cellData.strftime(date_format)
                        dataCol += str(cellData) + delimiter
                    if list_type is 'list':
                        dataCol.append(cellData)
            if skip_none is True and (
                (list_type is 'string' and dataCol =='')
                or
                (list_type is 'list' and len(dataCol) is 0)
                ):
                    continue
            dataList.append(dataCol)
        
        # Create single list if data calls for it
        if list_type is 'string' and len(dataList) == 1:
            dataList = str(dataList[0])
        if list_type is 'list':
            newList = []
            if col_range[0] + 1 == col_range[1]: 
                for d in dataList:
                    try:
                        newList.append(d[0])
                    except:
                        newList.append('')
                dataList = newList    
            if row_range[0] + 1 == row_range[1]:
                for d in dataList[0]:
                    try:
                        newList.append(d)
                    except:
                        newList.append('')
                dataList = newList
        if alerts is True:
            print('Done')
        return dataList
        
    def get_value(self, row=None, col=None):
        """ Return a single value of its original type """
        if row is None:
            row = self.row_range[0]
        if col is None:
            col = self.col_range[0]
        return self.sheet.cell(row=row, column=col).value

class convert:
    def __init__(self):
        # TODO various conversions from file types and such
        raise NotImplementedError